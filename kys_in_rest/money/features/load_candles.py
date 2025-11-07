import io
from datetime import datetime, timedelta, timezone
from statistics import median
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.worksheet.worksheet import Worksheet

from kys_in_rest.core.tg_utils import TgFeature
from kys_in_rest.money.features.repos.tinkoff_candles_repo import Candle, TinkoffCandlesRepo
from kys_in_rest.tg.entities.input_tg_msg import InputTgMsg
from kys_in_rest.tg.features.bot_msg_repo import BotMsgRepo
from kys_in_rest.users.features.check_admin import CheckTgAdmin


class LoadCandlesTgFeature(TgFeature):
    def __init__(
        self,
        bot_msg_repo: BotMsgRepo,
        check_tg_admin: CheckTgAdmin,
        candles_repo: TinkoffCandlesRepo,
    ):
        self.bot_msg_repo = bot_msg_repo
        self.check_tg_admin = check_tg_admin
        self.candles_repo = candles_repo

    async def do_async(self, msg: InputTgMsg) -> None:
        self.check_tg_admin.do(msg.tg_user_id)

        # Получаем тикер из сообщения
        ticker = (msg.text or "").strip().upper()
        if not ticker:
            await self.bot_msg_repo.send_text("Укажи тикер акции (например: NLMK)")
            return

        try:
            await self.bot_msg_repo.send_text(f"Загружаю свечи для {ticker}...")

            # Получаем свечи
            monthly_candles = self.candles_repo.get_monthly_candles(ticker, months=36)
            weekly_candles = self.candles_repo.get_weekly_candles(ticker)

            if not monthly_candles:
                await self.bot_msg_repo.send_text(f"Не удалось получить месячные свечи для {ticker}")
                return

            if not weekly_candles:
                await self.bot_msg_repo.send_text(f"Не удалось получить недельные свечи для {ticker}")
                return

            complete_monthly = self._filter_complete_months(monthly_candles)
            complete_weekly = self._filter_complete_weeks(weekly_candles)

            if not complete_monthly:
                await self.bot_msg_repo.send_text(f"Нет полных месячных свечей для {ticker}")
                return

            if not complete_weekly:
                await self.bot_msg_repo.send_text(f"Нет полных недельных свечей для {ticker}")
                return

            # Создаем Excel файл
            excel_bytes = self._create_excel(complete_monthly, complete_weekly, ticker)

            # Вычисляем статистику
            monthly_stats = self._calculate_statistics(complete_monthly)
            weekly_stats = self._calculate_statistics(complete_weekly)

            # Отправляем файл
            filename = f"{ticker}_candles.xlsx"
            await self.bot_msg_repo.send_document(
                document=excel_bytes,
                filename=filename,
                caption=f"Свечи {ticker}: месячные и недельные",
            )
            
            # Отправляем статистику
            stats_text = self._format_statistics(monthly_stats, weekly_stats, ticker)
            await self.bot_msg_repo.send_text(stats_text)
            
        except ValueError as e:
            await self.bot_msg_repo.send_text(f"Ошибка: {str(e)}")
        except Exception as e:
            await self.bot_msg_repo.send_text(f"Произошла ошибка: {str(e)}")

    def _create_excel(
        self,
        monthly_candles: list[Candle],
        weekly_candles: list[Candle],
        ticker: str,
    ) -> bytes:
        """Создает Excel файл со свечами"""
        wb = Workbook()
        month_sheet = wb.active
        if month_sheet is None:
            raise RuntimeError("Не удалось создать лист в Excel")
        self._fill_candles_sheet(
            sheet=month_sheet,
            title="Месяцы",
            candles=monthly_candles,
            date_format="%m.%Y",
        )

        week_sheet = wb.create_sheet(title="Недели")
        self._fill_candles_sheet(
            sheet=week_sheet,
            title="Недели",
            candles=weekly_candles,
            date_format="%d.%m.%Y",
        )

        # Сохраняем в байты
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        return excel_buffer.getvalue()

    def _fill_candles_sheet(
        self,
        *,
        sheet: Worksheet,
        title: str,
        candles: list[Candle],
        date_format: str,
    ) -> None:
        sheet.title = title

        # Заголовки
        sheet["A1"] = "Дата"
        sheet["B1"] = "Изменение (%)"

        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_font = Font(bold=True)
        sheet["A1"].fill = header_fill
        sheet["A1"].font = header_font
        sheet["B1"].fill = header_fill
        sheet["B1"].font = header_font

        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")

        for idx, candle in enumerate(sorted(candles, key=lambda c: c.time, reverse=True), start=2):
            sheet[f"A{idx}"] = candle.time.strftime(date_format)
            change_percent = ((candle.close - candle.open) / candle.open) * 100
            sheet[f"B{idx}"] = round(change_percent, 2)

            sheet[f"B{idx}"].fill = green_fill if change_percent >= 0 else red_fill

        sheet.column_dimensions["A"].width = 18
        sheet.column_dimensions["B"].width = 18

    def _calculate_statistics(self, candles: list[Candle]) -> dict[str, Any]:
        """Вычисляет статистику по свечам"""
        changes: list[float] = []
        for candle in candles:
            change_percent = ((candle.close - candle.open) / candle.open) * 100
            changes.append(change_percent)
        
        # Разделяем на рост и падение
        growths = [c for c in changes if c >= 0]
        falls = [c for c in changes if c < 0]
        
        stats: dict[str, Any] = {
            "total": len(changes),
            "growth_count": len(growths),
            "fall_count": len(falls),
        }
        
        if falls:
            stats["min_fall"] = round(max(falls), 2)
            stats["max_fall"] = round(min(falls), 2)
            stats["avg_fall"] = round(sum(falls) / len(falls), 2)
            stats["median_fall"] = round(median(falls), 2)
        else:
            stats["min_fall"] = None
            stats["max_fall"] = None
            stats["avg_fall"] = None
            stats["median_fall"] = None
        
        if growths:
            stats["min_growth"] = round(min(growths), 2)
            stats["max_growth"] = round(max(growths), 2)
            stats["avg_growth"] = round(sum(growths) / len(growths), 2)
            stats["median_growth"] = round(median(growths), 2)
        else:
            stats["min_growth"] = None
            stats["max_growth"] = None
            stats["avg_growth"] = None
            stats["median_growth"] = None

        stats["growth_probability"] = (
            round((stats["growth_count"] / stats["total"]) * 100)
            if stats["total"] > 0
            else 0
        )

        return stats

    def _format_statistics(
        self,
        monthly_stats: dict[str, Any],
        weekly_stats: dict[str, Any],
        ticker: str,
    ) -> str:
        """Форматирует статистику для отправки"""
        lines = [f"📊 {ticker}"]
        lines.append("")

        lines.append(self._format_period_stats("Месяцы", monthly_stats))
        lines.append(self._format_period_stats("Недели", weekly_stats))

        return "\n".join(lines)

    def _format_period_stats(self, label: str, stats: dict[str, Any]) -> str:
        red_part = (
            f"🔴 мин {self._fmt_percent(stats['min_fall'])}, "
            f"макс {self._fmt_percent(stats['max_fall'])}, "
            f"ср {self._fmt_percent(stats['avg_fall'])}, "
            f"мед {self._fmt_percent(stats['median_fall'])}"
        )

        green_part = (
            f"🟢 мин {self._fmt_percent(stats['min_growth'])}, "
            f"макс {self._fmt_percent(stats['max_growth'])}, "
            f"ср {self._fmt_percent(stats['avg_growth'])}, "
            f"мед {self._fmt_percent(stats['median_growth'])}"
        )

        growth_part = (
            f"📈 рост {stats['growth_probability']}% "
            f"({stats['growth_count']}/{stats['total']})"
        )

        return (
            f"{label} ({stats['total']}): {red_part} | "
            f"{green_part} | {growth_part}"
        )

    @staticmethod
    def _fmt_percent(value: float | None) -> str:
        if value is None:
            return "—"
        return f"{value:+.2f}%"

    def _filter_complete_months(self, candles: list[Candle]) -> list[Candle]:
        now = self._now_for_candles(candles).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        return [candle for candle in candles if candle.time < now]

    def _filter_complete_weeks(self, candles: list[Candle]) -> list[Candle]:
        now = self._now_for_candles(candles)
        week_start = (now - timedelta(days=now.weekday())).replace(
            hour=0,
            minute=0,
            second=0,
            microsecond=0,
        )
        return [candle for candle in candles if candle.time < week_start]

    @staticmethod
    def _now_for_candles(candles: list[Candle]) -> datetime:
        now = datetime.utcnow()
        if candles and candles[0].time.tzinfo is not None:
            return now.replace(tzinfo=timezone.utc)
        return now